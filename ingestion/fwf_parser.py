"""
src/ingestion/fwf_parser.py
============================
Step 1 of ingestion: Read official MoSPI layout + codes Excel files,
then parse raw fixed-width .txt microdata into a clean pandas DataFrame.

THE ONE RULE (never forget):
  MoSPI layout files use 1-based positions.
  Python slices are 0-based.
  Conversion:  python_start = official_start - 1   ← subtract 1
               python_end   = official_end           ← keep same

  Example: layout says Start=52, End=52 → Python: record[51:52]
"""

import io
import re
import logging
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Optional

log = logging.getLogger("statiq.ingestion.parser")


# ─────────────────────────────────────────────────────────────
# STEP 1A: PARSE LAYOUT EXCEL → COLSPECS
# ─────────────────────────────────────────────────────────────

def parse_layout_excel(xlsx_path: str) -> dict:
    """
    Parse official MoSPI data layout Excel (e.g. Datalayout_250_80R.xlsx).

    What it reads:
      Each row = one column in the raw .txt file.
      Columns: srl_no | Full_Name | Block | Item | Width | Byte_Start | ... | Field_Name

    Returns:
      {
        "lvl_01": {
          "meta":   { "record_length": 150, "filename": "H80_LVL_01.TXT" },
          "fields": [
            { "name": "rnd", "official_start": 1, "official_end": 2,
              "python_start": 0, "python_end": 2, "colspec": (0, 2) },
            ...
          ]
        },
        "lvl_02": { ... }
      }
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if 'CHHV1' in wb.sheetnames and 'CPERV1' in wb.sheetnames:
        # This is the PLFS layout file
        levels = {}
        for sheet_name, lvl_key, file_name, rec_len in [('CHHV1', 'lvl_01', 'CHHV1.TXT', 218), ('CPERV1', 'lvl_02', 'CPERV1.TXT', 371)]:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            fields = []
            cum_pos = 0
            for row in rows[1:]:
                full_name = row[1]
                field_len = row[4]
                byte_start = row[5]
                field_name = row[7]
                remarks = str(row[8] or '').strip()
                
                if not full_name or not field_len:
                    continue
                fn_str = str(full_name).strip()
                if fn_str.lower() in ('full name', 'nan', '', 'srl. no.') or fn_str.startswith('='):
                    continue
                    
                try:
                    field_len = int(float(str(field_len)))
                except (ValueError, TypeError):
                    continue
                    
                if isinstance(byte_start, (int, float)) and not isinstance(byte_start, bool):
                    start = int(byte_start)
                else:
                    start = cum_pos + 1
                    
                end = start + field_len - 1
                cum_pos = end
                
                name = str(field_name or full_name).strip()
                if name.startswith('=') or name.lower() == 'nan':
                    name = fn_str
                    
                fields.append({
                    'name':           name,
                    'full_name':      fn_str,
                    'length':         field_len,
                    'official_start': start,
                    'official_end':   end,
                    'python_start':   start - 1,
                    'python_end':     end,
                    'colspec':        (start - 1, end),
                    'remarks':        remarks,
                })
            levels[lvl_key] = {
                'meta': {
                    'record_length': rec_len,
                    'filename':      file_name,
                    'source_file':   Path(xlsx_path).name,
                },
                'fields': fields
            }
        log.info(f"[Parser] Layout parsed for PLFS: {len(levels)} levels")
        return levels

    # Find the data sheet (first sheet with layout/lvl/level in name, or sheet 0)
    target_sheet = wb.sheetnames[0]
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ['layout', 'lvl', 'level', 'data']):
            target_sheet = name
            break

    all_rows = list(wb[target_sheet].iter_rows(values_only=True))
    log.info(f"[Parser] Layout: {Path(xlsx_path).name} → sheet '{target_sheet}' "
             f"({len(all_rows)} rows)")
    return _rows_to_levels(all_rows, str(xlsx_path))


def _rows_to_levels(rows: list, source: str) -> dict:
    """
    Internal: walk through Excel rows, detect level headers,
    extract field specs, apply 1→0 index conversion.
    """
    levels   = {}
    cur_lvl  = None
    cur_meta = {}
    cur_flds = []
    cum_pos  = 0     # running byte position (for formula-based cells)

    for row in rows:
        c0 = str(row[0]) if row[0] is not None else ''

        # Detect new level/file header: "File: H80_LVL_01.TXT  RECORD LENTH:150+1"
        if ('File:' in c0 or 'FILE:' in c0) and 'TXT' in c0.upper():
            if cur_lvl and cur_flds:
                levels[cur_lvl] = {'meta': cur_meta, 'fields': cur_flds}

            m  = re.search(r'LVL[_\s]*(\d+)', c0, re.I)
            lm = re.search(r'LENTH[:\s]+(\d+)', c0, re.I)
            fm = re.search(r'([\w]+\.TXT)', c0, re.I)

            lvl_num  = m.group(1).zfill(2) if m else str(len(levels) + 1).zfill(2)
            cur_lvl  = f"lvl_{lvl_num}"
            cur_meta = {
                'record_length': int(lm.group(1)) if lm else None,
                'filename':      fm.group(1).upper() if fm else '',
                'source_file':   Path(source).name,
            }
            cur_flds = []
            cum_pos  = 0
            continue

        # If no level detected yet, try treating whole sheet as one level
        if cur_lvl is None:
            if row[1] and row[5]:
                fn = str(row[1]).strip()
                if fn and not fn.lower().startswith('full') and fn != 'nan':
                    cur_lvl  = 'lvl_01'
                    cur_meta = {'record_length': None, 'filename': '', 'source_file': Path(source).name}
                    cur_flds = []
                    cum_pos  = 0
        if not cur_lvl:
            continue

        # Extract field definition from this row
        # Col layout: 0=srl, 1=full_name, 2=block, 3=item, 4=sub_col,
        #             5=field_length, 6=byte_start, 7=dash, 8=byte_end,
        #             9=remarks, 10=field_name
        full_name  = row[1]
        field_len  = row[5]
        byte_start = row[6]
        field_name = row[10]
        remarks    = str(row[9] or '').strip()

        if not full_name or not field_len:
            continue
        fn_str = str(full_name).strip()
        if fn_str.lower() in ('full name', 'nan', '', 'srl. no.') or fn_str.startswith('='):
            continue

        try:
            field_len = int(float(str(field_len)))
        except (ValueError, TypeError):
            continue

        # Resolve byte start position
        # Layout Excel uses =I10+1 formulas; data_only=True may give None
        # → fall back to cumulative position tracker
        if isinstance(byte_start, (int, float)) and not isinstance(byte_start, bool):
            start = int(byte_start)
        else:
            start = cum_pos + 1

        end     = start + field_len - 1
        cum_pos = end   # advance tracker

        name = str(field_name or full_name).strip()
        if name.startswith('=') or name.lower() == 'nan':
            name = fn_str

        cur_flds.append({
            'name':           name,
            'full_name':      fn_str,
            'length':         field_len,
            'official_start': start,
            'official_end':   end,
            # ── THE 1→0 CONVERSION ────────────────────────────
            'python_start':   start - 1,     # subtract 1
            'python_end':     end,            # keep same
            'colspec':        (start - 1, end),
            # ─────────────────────────────────────────────────
            'remarks':        remarks,
        })

    if cur_lvl and cur_flds:
        levels[cur_lvl] = {'meta': cur_meta, 'fields': cur_flds}

    log.info(f"[Parser] Layout parsed: {len(levels)} levels")
    for lvl, d in levels.items():
        log.info(f"  {lvl}: {len(d['fields'])} fields, "
                 f"record_len={d['meta']['record_length']}, "
                 f"file={d['meta']['filename']}")
    return levels


# ─────────────────────────────────────────────────────────────
# STEP 1B: PARSE CODES EXCEL → CODEBOOK
# ─────────────────────────────────────────────────────────────

def parse_codes_excel(xlsx_path: str) -> dict:
    """
    Parse official MoSPI codes Excel (e.g. CODEs_for_Blocks_of_Sch__25_0.xlsx).

    What it reads:
      BLOCK | COL_ITEM | VARIABLE_NAME | CODE | CODE_DESCRIPTION

    Returns:
      { "variable_name_lower": { "code_string": "human_label" } }

    Note: codes are stored as strings to preserve leading zeros.
          "01" ≠ "1" for education codes. Always str.strip() before lookup.
    """
    wb       = openpyxl.load_workbook(xlsx_path)
    codebook = {}

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row to detect column positions
        header_row = rows[1] if len(rows) > 1 else rows[0]
        header     = [str(c or '').strip().lower() for c in header_row]

        var_col  = next((i for i, h in enumerate(header) if 'variable_name' in h or h == 'variable'), None)
        code_col = next((i for i, h in enumerate(header) if h == 'code'), None)
        desc_col = next((i for i, h in enumerate(header) if 'description' in h or 'label' in h), None)

        if var_col is None or code_col is None:
            continue

        start_row = 2  # skip title row + header row
        for row in rows[start_row:]:
            if not row or len(row) < max(v for v in [var_col, code_col, desc_col or 0] if v is not None) + 1:
                continue

            var  = str(row[var_col]  or '').strip()
            code = str(row[code_col] or '').strip()
            desc = str(row[desc_col] or '').strip() if desc_col is not None else ''

            if not var or var.lower() in ('nan', 'variable_name', ''):
                continue
            if not code or code.lower() in ('nan', 'code', 'none', ''):
                continue
            if not desc or desc.lower() in ('nan', 'none', 'code_description', ''):
                continue

            vk = var.lower()
            if vk not in codebook:
                codebook[vk] = {}
            codebook[vk][code] = desc.strip()

    log.info(f"[Parser] Codebook: {len(codebook)} variables, "
             f"{sum(len(v) for v in codebook.values())} codes")
    return codebook


def parse_state_codes(xlsx_path: str) -> dict:
    """
    Extract state code → state name from the 'State code' sheet.
    Returns: { "01": "Jammu & Kashmir", "33": "Tamil Nadu", ... }
    """
    wb     = openpyxl.load_workbook(xlsx_path)
    states = {}
    for sheet_name in wb.sheetnames:
        if 'state' in sheet_name.lower():
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                if row[0] and row[1]:
                    code = str(row[0]).strip()
                    name = str(row[1]).strip()
                    if code.isdigit():
                        states[code.zfill(2)] = name
    log.info(f"[Parser] State codes: {len(states)}")
    return states


# ─────────────────────────────────────────────────────────────
# STEP 2: READ FIXED-WIDTH FILE → DATAFRAME
# ─────────────────────────────────────────────────────────────

def read_fwf_level(
    data_path:  str,
    level_spec: dict,
    nrows:      Optional[int] = None,
) -> pd.DataFrame:
    """
    Read one NSS fixed-width .txt file using layout colspecs.

    Args:
        data_path:  path to raw .txt file (e.g. /data/raw/hces/h80_lvl_02.txt)
        level_spec: one level dict from parse_layout_excel()
        nrows:      row limit for testing (None = all rows)

    Returns:
        Raw DataFrame — all columns are strings before type casting.

    CRITICAL: encoding MUST be 'latin-1' for all MoSPI files.
              header=None because NSS files have NO header row.
    """
    fields   = level_spec['fields']
    colspecs = [f['colspec'] for f in fields]   # list of (python_start, python_end)
    names    = [f['name']    for f in fields]

    df = pd.read_fwf(
        data_path,
        colspecs  = colspecs,
        names     = names,
        encoding  = 'latin-1',   # ← ALWAYS latin-1 for MoSPI
        header    = None,        # ← NO header row in NSS files
        dtype     = str,         # ← read everything as string first
        na_values = [''],
        nrows     = nrows,
    )
    log.info(f"[Parser] Read: {Path(data_path).name} → {len(df):,} rows × {len(df.columns)} cols")
    return df


def read_fwf_from_bytes(
    raw_bytes:  bytes,
    level_spec: dict,
    nrows:      Optional[int] = None,
) -> pd.DataFrame:
    """
    Same as read_fwf_level but from in-memory bytes.
    Used when the file is downloaded from MinIO without writing to disk.
    """
    fields   = level_spec['fields']
    colspecs = [f['colspec'] for f in fields]
    names    = [f['name']    for f in fields]

    buffer = io.StringIO(raw_bytes.decode('latin-1'))
    df = pd.read_fwf(
        buffer,
        colspecs  = colspecs,
        names     = names,
        header    = None,
        dtype     = str,
        na_values = [''],
        nrows     = nrows,
    )
    return df


# ─────────────────────────────────────────────────────────────
# STEP 3: APPLY CODEBOOK LABELS
# ─────────────────────────────────────────────────────────────

def apply_codebook_labels(
    df:       pd.DataFrame,
    codebook: dict,
    suffix:   str = '_label',
) -> pd.DataFrame:
    """
    Add human-readable label columns next to every coded field.

    For each column in df that matches a codebook variable name,
    creates a new column: original_column + suffix (default: _label)

    Example:
      df['b3c4'] = ['1', '2', '1']
      → df['b3c4_label'] = ['male', 'female', 'male']

    Unknown codes → NaN (never raises an error).
    """
    added = []
    for col in df.columns:
        vk = col.lower()
        if vk in codebook:
            label_col    = col + suffix
            df[label_col] = df[col].astype(str).str.strip().map(codebook[vk])
            added.append(label_col)

    if added:
        log.debug(f"[Parser] Label columns added: {added}")
    return df


# ─────────────────────────────────────────────────────────────
# STEP 4: CAST TYPES AND CLEAN
# ─────────────────────────────────────────────────────────────

# Known numeric columns across all NSS surveys
NUMERIC_COLS = {
    # HCES Health
    'hhsz','b3c5','b3c10','b4c4','b4c7',
    'b6i3','b6i4','b6i12','b6i20','b6i24',
    'b7i6','b7i7','b7i8','b7i9','b7i10','b7i11','b7i12',
    'b7i13','b7i14','b7i15','b7i16','b7i20',
    'b8i3','b8i4','b8i8',
    'b9i11','b9i12','b9i13','b9i14','b9i15','b9i16',
    'b9i17','b9i18','b9i19','b9i20','b9i24',
    'b10i3','b10i7',
    'b5i6','b5i7','b5i8','b5i9','b5i10','b5i11',
    'umce','mult','nst','nstj','subdvsn','caph','smah',
    # PLFS
    'age','multiplier','census_count',
}


def cast_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert numeric fields to float/int and clean string fields.
    Auto-detects numeric columns from known list + name patterns.
    """
    for col in df.columns:
        if col.endswith('_label'):
            continue

        is_numeric = (
            col.lower() in NUMERIC_COLS
            or col.lower().startswith('mult')
        )
        if is_numeric:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        else:
            # Strip whitespace, replace string 'nan' with actual NaN
            df[col] = (
                df[col].astype(str).str.strip()
                    .replace({'nan': pd.NA, 'None': pd.NA, '': pd.NA})
            )
    return df


# ─────────────────────────────────────────────────────────────
# STEP 5: DATA QUALITY VALIDATION
# ─────────────────────────────────────────────────────────────

def validate_dataframe(df: pd.DataFrame, survey: str) -> dict:
    """
    Run basic data quality checks. Returns a QC report dict.

    Checks:
      - Null state code
      - Null gender
      - Age outliers (< 0 or > 120)
      - Invalid activity codes (PLFS)

    Severity: OK | WARN (< 1% of rows) | ERROR (≥ 1% of rows)
    """
    VALID_ACTIVITIES = {
        '11','12','21','22','31','41','51',
        '61','71','81','91','92','93','94','95','97','98','99'
    }

    report = {'survey': survey, 'total_rows': len(df), 'issues': []}
    checks = {}

    # Check state
    st_col = next((c for c in ['st', 'State', 'state_code'] if c in df.columns), None)
    if st_col:
        checks['null_state'] = int(df[st_col].isna().sum())

    # Check gender
    g_col = next((c for c in ['b3c4', 'gender', 'Gender'] if c in df.columns), None)
    if g_col:
        checks['null_gender'] = int(df[g_col].isna().sum())

    # Check age
    a_col = next((c for c in ['b3c5', 'age', 'Age'] if c in df.columns), None)
    if a_col:
        age = pd.to_numeric(df[a_col], errors='coerce')
        checks['null_age']    = int(age.isna().sum())
        checks['age_outliers']= int(((age < 0) | (age > 120)).sum())

    # Check activity codes (PLFS)
    act_col = next((c for c in ['usual_activity', 'Usual_Activity'] if c in df.columns), None)
    if act_col:
        invalid = ~df[act_col].astype(str).str.strip().isin(VALID_ACTIVITIES | {'nan', ''})
        checks['invalid_activity'] = int(invalid.sum())

    for check, count in checks.items():
        pct    = count / max(len(df), 1) * 100
        status = 'OK' if count == 0 else ('WARN' if pct < 1.0 else 'ERROR')
        level  = log.info if status == 'OK' else log.warning
        level(f"  QC [{status:5s}] {check}: {count:,} rows ({pct:.2f}%)")
        if count > 0:
            report['issues'].append({
                'check': check, 'count': count, 'pct': round(pct, 3), 'status': status
            })

    return report
